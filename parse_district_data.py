"""Parse the AP district-wise GVA/GSDP/GDDP workbook into RAG-friendly text chunks.

Source: district x sector blocks, 4 years each (2022-23 TRE, 2023-24 SRE, 2024-25 FRE,
2025-26 FAE), with Value (Rs. Cr.), Rank, Growth %, Contribution % per year.

Output:
  - structured_district_data.csv  — full long-format table (exact-number source of truth)
  - corpus_files/District_Data/<District>/<Sector>.txt  — one small chunk per district-sector
  - corpus_files/District_Data/<District>_Snapshot.txt   — comparative-advantage summary per district
"""
import csv
import os

import openpyxl

XLSX_PATH = "/Users/thesinghaa/Downloads/1 DB 08-03-2026.xlsx"
BASE_DIR = os.path.dirname(__file__)
CSV_OUT = os.path.join(BASE_DIR, "structured_district_data.csv")
DISTRICT_DATA_DIR = os.path.join(BASE_DIR, "corpus_files", "District_Data")

YEARS = ["2022-23 (TRE)", "2023-24 (SRE)", "2024-25 (FRE)", "2025-26 (FAE)"]
# column offsets (0-indexed from column D=col 4) per year: (value, rank, growth_or_None, contri)
YEAR_COLS = [
    (4, 5, None, 6),      # 2022-23: D,E,F  (no growth — first year)
    (7, 8, 9, 10),         # 2023-24: G,H,I,J
    (11, 12, 13, 14),      # 2024-25: K,L,M,N
    (15, 16, 17, 18),      # 2025-26: O,P,Q,R
]

# Sectors that are aggregates/derived rows, not raw economic sectors — still useful for
# GDDP/NDDP/population/per-capita lookups, so we keep them but tag them distinctly.
AGGREGATE_LABELS = {
    "AGRICULTURE \n&\nALLIED\n SECTOR": "Agriculture & Allied Sector (aggregate)",
    "INDUSTRY SECTOR": "Industry Sector (aggregate)",
    "SERVICES SECTOR": "Services Sector (aggregate)",
    "GDVA": "Gross District Value Added (GDVA)",
    "PRODUCT TAXES": "Product Taxes",
    "PRODUCT SUBSIDIES": "Product Subsidies",
    "GDDP": "Gross District Domestic Product (GDDP)",
    "NDDP": "Net District Domestic Product (NDDP)",
    "POPULATION(‘000)": "Population ('000)",
    "PER CAPITA IN Rs.": "Per Capita Income (Rs.)",
}


def clean(s):
    return " ".join(str(s).split()) if s is not None else s


def load_blocks(ws):
    """Yield (sector_name, {district: {year_label: {value,rank,growth,contri}}})."""
    current_sector = None
    current_data = {}
    row = 4
    max_row = ws.max_row
    while row <= max_row:
        b = ws.cell(row=row, column=2).value
        c = ws.cell(row=row, column=3).value
        if b:  # new sector block starts
            if current_sector and current_data:
                yield current_sector, current_data
            current_sector = clean(b)
            current_data = {}
        if c and clean(c) != "Total":
            district = clean(c)
            years = {}
            for yi, (vcol, rcol, gcol, ccol) in enumerate(YEAR_COLS):
                val = ws.cell(row=row, column=vcol).value
                rank = ws.cell(row=row, column=rcol).value
                growth = ws.cell(row=row, column=gcol).value if gcol else None
                contri = ws.cell(row=row, column=ccol).value
                years[YEARS[yi]] = {
                    "value": val,
                    "rank": rank,
                    "growth_pct": growth,
                    "contribution_pct": contri,
                }
            current_data[district] = years
        row += 1
    if current_sector and current_data:
        yield current_sector, current_data


def fmt_num(x, decimals=2):
    if x is None:
        return "N/A"
    try:
        return f"{float(x):,.{decimals}f}"
    except (TypeError, ValueError):
        return str(x)


def sector_display_name(raw):
    return AGGREGATE_LABELS.get(raw, raw)


def write_district_sector_chunk(district, sector, years_data):
    sector_label = sector_display_name(sector)
    safe_sector = "".join(c if c.isalnum() or c in " -_&" else "_" for c in sector_label).strip()
    dist_dir = os.path.join(DISTRICT_DATA_DIR, district.title().replace(" ", "_"))
    os.makedirs(dist_dir, exist_ok=True)
    path = os.path.join(dist_dir, f"{safe_sector}.txt")

    lines = [f"District: {district.title()} — Sector: {sector_label}", ""]
    for year in YEARS:
        d = years_data[year]
        parts = [f"Value = Rs. {fmt_num(d['value'])} Cr.", f"Rank among 28 districts = {d['rank']}"]
        if d["growth_pct"] is not None:
            parts.append(f"YoY Growth = {fmt_num(d['growth_pct'])}%")
        parts.append(f"Contribution to state total = {fmt_num(d['contribution_pct'])}%")
        lines.append(f"{year}: " + ", ".join(parts))
    with open(path, "w") as f:
        f.write("\n".join(lines))


def write_district_snapshot(district, sector_rows):
    """sector_rows: list of (sector_label, latest_year_dict, is_aggregate)."""
    latest_year = YEARS[-1]
    non_aggregate = [r for r in sector_rows if not r[2]]
    by_contribution = sorted(
        non_aggregate, key=lambda r: (r[1]["contribution_pct"] or 0), reverse=True
    )
    by_growth = sorted(
        [r for r in non_aggregate if r[1]["growth_pct"] is not None],
        key=lambda r: r[1]["growth_pct"],
        reverse=True,
    )
    by_rank = sorted(non_aggregate, key=lambda r: (r[1]["rank"] or 999))

    lines = [f"District Economic Profile Snapshot: {district.title()} (latest year: {latest_year})", ""]

    lines.append("Top sectors by contribution to district GVA (comparative advantage):")
    for label, d, _ in by_contribution[:5]:
        lines.append(
            f"  - {label}: {fmt_num(d['contribution_pct'])}% of district GVA, "
            f"Rs. {fmt_num(d['value'])} Cr., statewide rank {d['rank']}"
        )
    lines.append("")

    lines.append(f"Fastest-growing sectors in {district.title()} ({latest_year} YoY):")
    for label, d, _ in by_growth[:5]:
        lines.append(f"  - {label}: {fmt_num(d['growth_pct'])}% growth, Rs. {fmt_num(d['value'])} Cr.")
    lines.append("")

    lines.append(f"Best statewide ranks (sectors where {district.title()} leads other districts):")
    for label, d, _ in by_rank[:5]:
        lines.append(f"  - {label}: rank {d['rank']} of 28, Rs. {fmt_num(d['value'])} Cr.")
    lines.append("")

    aggregates = [r for r in sector_rows if r[2]]
    lines.append("Aggregate figures:")
    for label, d, _ in aggregates:
        g = f", growth {fmt_num(d['growth_pct'])}%" if d["growth_pct"] is not None else ""
        lines.append(f"  - {label}: Rs. {fmt_num(d['value'])} Cr.{g}, rank {d['rank']}")

    path = os.path.join(DISTRICT_DATA_DIR, f"{district.title().replace(' ', '_')}_Snapshot.txt")
    with open(path, "w") as f:
        f.write("\n".join(lines))


def main():
    os.makedirs(DISTRICT_DATA_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]

    csv_rows = []
    # district -> list of (sector_label, latest_year_data, is_aggregate)
    district_sector_latest = {}

    for sector, districts in load_blocks(ws):
        is_aggregate = sector in AGGREGATE_LABELS
        sector_label = sector_display_name(sector)
        for district, years in districts.items():
            write_district_sector_chunk(district, sector, years)
            district_sector_latest.setdefault(district, []).append(
                (sector_label, years[YEARS[-1]], is_aggregate)
            )
            for year, d in years.items():
                csv_rows.append(
                    {
                        "district": district,
                        "sector": sector_label,
                        "year": year,
                        "value_rs_cr": d["value"],
                        "rank": d["rank"],
                        "growth_pct": d["growth_pct"],
                        "contribution_pct": d["contribution_pct"],
                    }
                )

    for district, rows in district_sector_latest.items():
        write_district_snapshot(district, rows)

    with open(CSV_OUT, "w", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["district", "sector", "year", "value_rs_cr", "rank", "growth_pct", "contribution_pct"],
        )
        writer.writeheader()
        writer.writerows(csv_rows)

    n_districts = len(district_sector_latest)
    n_files = sum(len(files) for _, _, files in os.walk(DISTRICT_DATA_DIR))
    print(f"Parsed {n_districts} districts, wrote {n_files} files under {DISTRICT_DATA_DIR}")
    print(f"Long-format CSV: {CSV_OUT} ({len(csv_rows)} rows)")


if __name__ == "__main__":
    main()
